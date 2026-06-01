"""
xlsx 内置技能 — Excel 表格 (.xlsx/.xlsm/.csv/.tsv) 的创建、阅读和编辑。
支持公式、数据分析、多工作表操作。
"""
from pathlib import Path
from typing import Any, Optional
from loguru import logger

SKILL_NAME = "xlsx"
SKILL_DESCRIPTION = "创建、读取和编辑表格文件 (.xlsx/.csv/.tsv)，支持多工作表、公式和数据分析"

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


async def execute(
    action: str,
    file_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
    max_rows: int = 100,
    **kwargs: Any,
) -> dict[str, Any]:
    """
    执行表格操作。

    Args:
        action: 操作类型（read/info/sheets）
        file_path: 文件路径
        sheet_name: 工作表名称
        max_rows: 最大读取行数

    Returns:
        操作结果
    """
    if not HAS_OPENPYXL:
        return {
            "success": False,
            "error": "缺少 openpyxl 依赖，请运行: pip install openpyxl",
        }

    if action not in {"read", "info", "sheets"}:
        return {"success": False, "error": f"不支持的操作: {action}"}

    if not file_path:
        return {"success": False, "error": "需要提供 file_path"}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path.cwd() / path
    if not path.exists():
        return {"success": False, "error": f"文件不存在: {file_path}"}

    logger.bind(event="xlsx_skill", action=action).info("表格操作")

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

        if action == "sheets":
            return {
                "success": True,
                "file": str(path),
                "sheets": wb.sheetnames,
                "active_sheet": wb.active.title if wb.active else None,
            }

        elif action == "info":
            sheet = wb[sheet_name] if sheet_name else wb.active
            return {
                "success": True,
                "file": str(path),
                "sheet": sheet.title,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "sheets": wb.sheetnames,
            }

        elif action == "read":
            sheet = wb[sheet_name] if sheet_name else wb.active
            rows = []
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= max_rows:
                    break
                rows.append([str(c) if c is not None else "" for c in row])

            wb.close()
            return {
                "success": True,
                "file": str(path),
                "sheet": sheet.title,
                "rows_read": len(rows),
                "total_rows": sheet.max_row,
                "total_columns": sheet.max_column,
                "headers": rows[0] if rows else [],
                "data": rows[1:11] if len(rows) > 1 else [],
                "has_more": len(rows) < sheet.max_row,
            }

        wb.close()
    except Exception as e:
        return {"success": False, "error": f"表格操作失败: {str(e)}"}

    return {"success": False}
